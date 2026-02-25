import os
import re
from collections import defaultdict
from fpdf import FPDF
import openpyxl
from datetime import datetime
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

# ============================================================
# CONFIGURAÇÕES - AJUSTE AQUI
# ============================================================
ARQUIVO_EXCEL = r"C:\Documentos\Projects\pyCharm\gera-relatorios\CBMDF\atendimentos-cbmdf.xlsx"
LOGO_PATH = r"C:\Documentos\Projects\pyCharm\gera-relatorios\logo_PNG_SEMFUNDO_01.png"
TIMBRADO_PATH = r"C:\Documentos\Projects\pyCharm\gera-relatorios\papaeltimbrado.png"
PASTA_SAIDA = r"C:\Documentos\Projects\pyCharm\gera-relatorios\relatorios"

# Nomes das colunas na planilha (Plano e Tipo Atendimento invertidos na planilha)
COL_PACIENTE = "Paciente"
COL_PLANO = "Tipo Atendimento"
COL_TIPO_ATENDIMENTO = "Plano"
COL_DATA = "Data"
COL_STATUS = "Status"
COL_TIPO_FILIAL = "Tipo Filial"

# Meses por extenso
MESES = {
    1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril",
    5: "maio", 6: "junho", 7: "julho", 8: "agosto",
    9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro",
}

# Data do documento gerada automaticamente (data de hoje)
hoje = datetime.now()
DATA_DOCUMENTO = f"{hoje.day} de {MESES[hoje.month]} de {hoje.year}"

# Diagnóstico fixo para ABA
DIAGNOSTICO = "TRANSTORNO DO ESPECTRO DO AUTISMO (CID F84.0)"

# Mapeamento: Tipo Atendimento -> nome que aparece no PDF
MAPA_ESPECIALIDADE = {
    "TERAPIA ABA - SESSAO": "PSICOLOGIA (TERAPIA ABA)",
    "TERAPIA ABA - ATENDIMENTO SEMANAL CONFORME ESPECIFICACAO MEDICA": "PSICOLOGIA (TERAPIA ABA)",
    "PSICOTERAPIA INDIVIDUAL": "PSICOLOGIA",
    "AVALIACAO PSICOLOGICA": "PSICOLOGIA",
    "PSICOPEDAGOGIA INDIVIDUAL": "PSICOPEDAGOGIA",
    "PSICOMOTRICIDADE INDIVIDUAL": "PSICOMOTRICIDADE",
    "SESSOES DE FONOTERAPIA/FONOAUDIOLOGIA": "FONOAUDIOLOGIA",
    "TERAPIA OCUPACIONAL - AVALIACAO DOS COMPONENTES DE DESEMPENHO OCUPACIONAL - SESSOES": "TERAPIA OCUPACIONAL",
}

# Números por extenso
EXTENSO = {
    1: "uma", 2: "duas", 3: "três", 4: "quatro", 5: "cinco",
    6: "seis", 7: "sete", 8: "oito", 9: "nove", 10: "dez",
    11: "onze", 12: "doze", 13: "treze", 14: "quatorze", 15: "quinze",
    16: "dezesseis", 17: "dezessete", 18: "dezoito", 19: "dezenove", 20: "vinte",
    21: "vinte e uma", 22: "vinte e duas", 23: "vinte e três", 24: "vinte e quatro",
    25: "vinte e cinco", 26: "vinte e seis", 27: "vinte e sete", 28: "vinte e oito",
    29: "vinte e nove", 30: "trinta", 31: "trinta e uma", 32: "trinta e duas",
    33: "trinta e três", 34: "trinta e quatro", 35: "trinta e cinco", 36: "trinta e seis",
    37: "trinta e sete", 38: "trinta e oito", 39: "trinta e nove", 40: "quarenta",
    41: "quarenta e uma", 42: "quarenta e duas", 43: "quarenta e três",
    44: "quarenta e quatro", 45: "quarenta e cinco", 46: "quarenta e seis",
    47: "quarenta e sete", 48: "quarenta e oito", 49: "quarenta e nove",
    50: "cinquenta", 51: "cinquenta e uma", 52: "cinquenta e duas",
    53: "cinquenta e três", 54: "cinquenta e quatro", 55: "cinquenta e cinco",
    56: "cinquenta e seis", 57: "cinquenta e sete", 58: "cinquenta e oito",
    59: "cinquenta e nove", 60: "sessenta",
}

# Textos fixos da Proposta de Intervenção (ABA)
TEXTO_PROPOSTA_INTERVENCAO = (
    "A educação no mundo contemporâneo favorece o desenvolvimento de habilidades e "
    "competências das crianças e jovens, com vistas não apenas à formação acadêmica, "
    "mas ao desenvolvimento dos aspectos socioemocionais e, sobretudo, na construção "
    "de um projeto de vida, exigindo assim um olhar coletivo sobre o indivíduo. Por isso, "
    "tanto a legislação educacional - como, por exemplo, a Base Nacional Curricular "
    "Comum (nova BNCC) - quanto às abordagens psicopedagógicas atuais - como as "
    "inteligências múltiplas de Gardner ou alguns achados neurocientíficos - abriram "
    "novas perspectivas para o processo de ensino e aprendizagem."
)

TEXTO_PROPOSTA_INTERVENCAO_2 = (
    "A multiplicidade interventiva aproxima o paciente da realidade socioemocional de "
    "uma maneira leve e espontânea, possibilitando desenvolver suas habilidades e "
    "competências para o mundo social que compartilhamos atualmente."
)

TEXTO_PROPOSTA_INTERVENCAO_3 = (
    "A ciência ABA tem como foco o trabalho de estimulação, aquisição de novas "
    "habilidades, ampliação, remodelação e reforço comportamental no âmbito social, "
    "comunicativo, cognitivo, emocional e acadêmico. As intervenções devem ser feitas de "
    "forma contínua e repetitivas, para aumento dos comportamentos e habilidades. Essa "
    "abordagem consiste em conjunto a terapia psicológica, cujo objetivo é observar as "
    "habilidades atencionais, sociais, cognitivas, possíveis déficits e dificuldades, assim "
    "como aspectos psicológicos."
)


# ============================================================
# CLASSE DO PDF
# ============================================================
class RelatorioPDF(FPDF):
    def __init__(self, logo_path=None, timbrado_path=None , logo_only_first=False, logo_right=False):
        super().__init__()
        self.logo_path = logo_path
        self.timbrado_path = timbrado_path
        self.logo_only_first = logo_only_first
        self.logo_right = logo_right
        # Margens ABNT: 3cm esquerda/superior, 2cm direita/inferior
        self.set_left_margin(30)
        self.set_right_margin(20)
        self.set_top_margin(30)

    def header(self):
        # 1. Papel timbrado como fundo (todas as páginas)
        if self.timbrado_path and os.path.exists(self.timbrado_path):
            self.image(self.timbrado_path, x=0, y=0, w=210, h=297)
        # Logo Colorida
        if self.logo_path and os.path.exists(self.logo_path):
            if self.logo_only_first and self.page_no() > 1:
                self.ln(10)
                return
            logo_w = 45
            if self.logo_right:
                x_pos = 210 - 20 - logo_w
            else:
                x_pos = (210 - logo_w) / 2
            self.image(self.logo_path, x=x_pos, y=10, w=logo_w)
            self.ln(35)
        else:
            self.ln(10)

    def footer(self):
        pass


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================
def write_mixed(pdf, parts):
    """Escreve texto com trechos alternando negrito/normal, sem cortar palavras."""
    md_text = ""
    for part in parts:
        if part.get("bold"):
            md_text += f"**{part['text']}**"
        else:
            md_text += part["text"]
    pdf.set_font("Helvetica", "", 12)
    pdf.multi_cell(0, 7, md_text, markdown=True)


def sanitize(text):
    """Remove caracteres problemáticos para nomes de arquivo."""
    text = text.strip()
    text = re.sub(r'[\\/*?:"<>|]', "", text)
    text = text.replace(" ", "_")
    return text


def get_convenio_info(plano, is_aba=False):
    """Retorna (saude, nome_completo) baseado no plano."""
    p = plano.strip().upper()
    if p == "CBMDF":
        return (
            "DIRETORIA DE SAÚDE",
            "CORPO DE BOMBEIROS MILITAR DO DISTRITO FEDERAL (CBMDF)",
        )
    elif p == "PMDF":
        return (
            "SAÚDE PMDF",
            "POLÍCIA MILITAR DO DISTRITO FEDERAL (PMDF)",
        )
    elif p == "FUSEX":
        return (
            "FUNDO DE SAÚDE DO EXÉRCITO (FUSEX)",
            "HOSPITAL MILITAR DA ÁREA DE BRASÍLIA (HMAB)",
        )
    else:
        return (f"SAÚDE {p}", p)


# ============================================================
# GERADOR PDF - MODELO TÍPICO (Psicoterapia Individual)
# ============================================================
def gerar_pdf_tipico(nome, plano, especialidade_label, sessoes, mes_ref, ano_ref, filial="Matriz"):
    """Gera PDF modelo Típico (1 página) - logo à direita."""

    pdf = RelatorioPDF(logo_path=LOGO_PATH, timbrado_path=TIMBRADO_PATH, logo_only_first=False, logo_right=True)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=35)

    # --- Data ---
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, f"Brasília, {DATA_DOCUMENTO}", 0, 1, "R")
    pdf.ln(10)

    # --- Destinatário ---
    saude, nome_completo = get_convenio_info(plano, is_aba=False)

    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 6, "Ao", 0, 1, "L")
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 6, saude, 0, 1, "L")
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 6, nome_completo, 0, 1, "L")
    pdf.set_font("Helvetica", "U", 12)
    pdf.cell(0, 6, "BRASÍLIA - DF", 0, 1, "L")
    pdf.ln(12)

    # --- Saudação ---
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, "Prezados(as) Senhores(as)", 0, 1, "L")
    pdf.ln(9)

    # --- Parágrafo único ---
    sessoes_int = int(sessoes)
    extenso = EXTENSO.get(sessoes_int, str(sessoes_int))
    mes_nome = MESES.get(mes_ref, str(mes_ref))
    nome_upper = nome.strip().upper()

    write_mixed(pdf, [
        {"text": "Solicitamos autorização para realização de "},
        {"text": f"{sessoes_int} ({extenso}) ", "bold": True},
        {"text": "sessões de "},
        {"text": especialidade_label, "bold": True},
        {"text": " para o(a) paciente "},
        {"text": nome_upper, "bold": True},
        {"text": ", para o mês de "},
        {"text": f"{mes_nome} de {ano_ref}.", "bold": True},
        {"text": " O(a) paciente necessita de acompanhamento constante na especialidade "
                 "mencionada para obtenção de bom resultado terapêutico."},
    ])
    pdf.ln(10)

    # --- Despedida ---
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, "Atenciosamente,", 0, 1, "L")

    # --- Salvar ---
    pasta_filial = os.path.join(PASTA_SAIDA, filial)
    os.makedirs(pasta_filial, exist_ok=True)
    nome_arquivo = f"{sanitize(nome)}_{sanitize(especialidade_label)}_TIPICO.pdf"
    caminho = os.path.join(pasta_filial, nome_arquivo)
    pdf.output(caminho)
    return caminho


# ============================================================
# GERADOR PDF - MODELO ABA (2 páginas)
# ============================================================
def gerar_pdf_aba(nome, plano, sessoes, mes_ref, ano_ref, filial="Matriz"):
    """Gera PDF modelo ABA (2 páginas) - logo à direita, só na 1ª página."""

    pdf = RelatorioPDF(logo_path=LOGO_PATH, timbrado_path=TIMBRADO_PATH, logo_only_first=True, logo_right=True)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=35)

    # --- Data ---
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, f"Brasília, {DATA_DOCUMENTO}", 0, 1, "R")
    pdf.ln(10)

    # --- Destinatário ---
    saude, nome_completo = get_convenio_info(plano, is_aba=True)

    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 6, "Ao", 0, 1, "L")
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 6, saude, 0, 1, "L")
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 6, nome_completo, 0, 1, "L")
    pdf.set_font("Helvetica", "U", 12)
    pdf.cell(0, 6, "BRASÍLIA - DF", 0, 1, "L")
    pdf.ln(12)

    # --- Saudação ---
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, "Prezados(as) Senhores(as)", 0, 1, "L")
    pdf.ln(9)

    # --- Parágrafo introdutório ---
    nome_upper = nome.strip().upper()
    write_mixed(pdf, [
        {"text": "Informamos que o(a) paciente "},
        {"text": nome_upper, "bold": True},
        {"text": " foi encaminhado a esta clínica, por essa diretoria"},
        {"text": ", para atendimento em terapias multi e interdisciplinares, com uso da ciência "},
        {"text": "ABA", "bold": True},
        {"text": " (Applied Behavior Analysis), por tempo indeterminado."},
    ])
    pdf.ln(8)

    # --- PROPOSTA DE INTERVENÇÃO ---
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "PROPOSTA DE INTERVENÇÃO", 0, 1, "C")
    pdf.ln(6)

    pdf.set_font("Helvetica", "", 12)
    pdf.multi_cell(0, 7, TEXTO_PROPOSTA_INTERVENCAO)
    pdf.ln(6)
    pdf.multi_cell(0, 7, TEXTO_PROPOSTA_INTERVENCAO_2)
    pdf.ln(6)
    pdf.multi_cell(0, 7, TEXTO_PROPOSTA_INTERVENCAO_3)
    pdf.ln(8)

    # --- PROPOSTA DE ATENDIMENTO ---
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "PROPOSTA DE ATENDIMENTO", 0, 1, "C")
    pdf.ln(6)

    sessoes_int = int(sessoes)
    extenso = EXTENSO.get(sessoes_int, str(sessoes_int))
    mes_nome = MESES.get(mes_ref, str(mes_ref))

    write_mixed(pdf, [
        {"text": "Com vistas à implantação da intervenção proposta, solicitamos autorização para "
                 "realização de "},
        {"text": f"{sessoes_int} ({extenso})", "bold": True},
        {"text": " sessões de "},
        {"text": "TERAPIA ABA", "bold": True},
        {"text": " no mês de "},
        {"text": f"{mes_nome} de {ano_ref}.", "bold": True},
    ])
    pdf.ln(12)

    # --- Despedida ---
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, "Atenciosamente,", 0, 1, "L")

    # --- Salvar ---
    pasta_filial = os.path.join(PASTA_SAIDA, filial)
    os.makedirs(pasta_filial, exist_ok=True)
    nome_arquivo = f"{sanitize(nome)}_TERAPIA_ABA.pdf"
    caminho = os.path.join(pasta_filial, nome_arquivo)
    pdf.output(caminho)
    return caminho


# ============================================================
# LEITURA E AGRUPAMENTO
# ============================================================
def main():
    wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
    ws = wb.active

    # Lê cabeçalhos
    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    # Para mostrar as colunas e os indices da planilha
    #print(f"Colunas encontradas: {headers}")
    #print(f"Índices: {idx}")
    #print()

    # Primeiro passo: coleta todas as sessões por paciente
    # paciente_sessoes: (paciente, plano, filial) -> lista de (tipo_atend, especialidade_profissional)
    paciente_sessoes = defaultdict(list)
    meses_encontrados = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        paciente = row[idx[COL_PACIENTE]]
        plano = row[idx[COL_PLANO]]
        tipo_atend = row[idx[COL_TIPO_ATENDIMENTO]]
        data = row[idx[COL_DATA]]
        filial = row[idx[COL_TIPO_FILIAL]]
        especialidade_prof = row[idx["Especialidade"]] if "Especialidade" in idx else ""

        if not paciente or not tipo_atend:
            continue

        # Extrai mês/ano da data
        if hasattr(data, "month"):
            mes, ano = data.month, data.year
        else:
            partes = str(data).split("/")
            mes, ano = int(partes[1]), int(partes[2])

        meses_encontrados.add((mes, ano))

        tipo_upper = tipo_atend.strip().upper()
        filial_nome = filial.strip() if filial else "Matriz"
        esp_prof = especialidade_prof.strip() if especialidade_prof else ""
        chave = (paciente.strip(), plano.strip(), filial_nome)
        paciente_sessoes[chave].append((tipo_upper, esp_prof))

    # Segundo passo: decide ABA ou Típico por paciente
    # Regras:
    #   1. Se o paciente tem mais de 5 sessões -> ABA
    #   2. Se tem até 5 sessões:
    #      - Se TODOS os profissionais são "Psicólogo Clínico" -> Típico
    #      - Se QUALQUER profissional é outra especialidade -> ABA
    aba_contagem = defaultdict(int)
    tipico_contagem = defaultdict(int)

    for chave, sessoes_lista in paciente_sessoes.items():
        total_sessoes = len(sessoes_lista)
        especialidades_prof = [esp for _, esp in sessoes_lista]

        if total_sessoes > 5:
            # Mais de 5 sessões -> sempre ABA
            aba_contagem[chave] = total_sessoes
        else:
            # Até 5 sessões -> verificar especialidade do profissional
            todos_psicologo_clinico = all(
                esp.upper() == "PSICÓLOGO CLÍNICO" or esp.upper() == "PSICOLOGO CLINICO"
                for esp in especialidades_prof
                if esp  # ignora vazios
            )

            if todos_psicologo_clinico and especialidades_prof:
                # Todos são Psicólogo Clínico -> Típico
                especialidade_label = "PSICOLOGIA"
                chave_tipico = (chave[0], chave[1], especialidade_label, chave[2])
                tipico_contagem[chave_tipico] = total_sessoes
            else:
                # Algum profissional é ABA/Fono/Fisio/etc -> ABA
                aba_contagem[chave] = total_sessoes

    # Determina mês/ano de referência
    if meses_encontrados:
        mes_ref, ano_ref = max(meses_encontrados)
    else:
        mes_ref, ano_ref = 1, 2026

    print(f"Mês de referência detectado: {MESES[mes_ref]}/{ano_ref}")
    print(f"Pacientes ABA: {len(aba_contagem)}")
    print(f"Pacientes Típico: {len(tipico_contagem)}")
    print(f"{'='*70}\n")

    total = 0

    # Gera PDFs ABA
    print("--- RELATÓRIOS ABA ---")
    for (paciente, plano, filial), sessoes in sorted(aba_contagem.items()):
        try:
            caminho = gerar_pdf_aba(paciente, plano, sessoes, mes_ref, ano_ref, filial)
            print(f"✓ {paciente:<40} | TERAPIA ABA          | {sessoes:>2} sessões | {filial}")
        except Exception as e:
            print(f"✗ ERRO em {paciente}: {e}")
        total += 1

    print()

    # Gera PDFs Típico
    print("--- RELATÓRIOS TÍPICO ---")
    for (paciente, plano, especialidade, filial), sessoes in sorted(tipico_contagem.items()):
        try:
            caminho = gerar_pdf_tipico(paciente, plano, especialidade, sessoes, mes_ref, ano_ref, filial)
            print(f"✓ {paciente:<40} | {especialidade:<20} | {sessoes:>2} sessões | {filial}")
        except Exception as e:
            print(f"✗ ERRO em {paciente}: {e}")
        total += 1

    print(f"\n{'='*70}")
    print(f"Total de relatórios gerados: {total}")
    print(f"Pasta de saída: {os.path.abspath(PASTA_SAIDA)}")


if __name__ == "__main__":
    main()